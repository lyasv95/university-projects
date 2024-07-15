from py2neo import Graph
import timeit
import statistics
import scipy.stats as stats
import openpyxl

def query1(user_id, stampa):
    result = tx.run("MATCH (u:Utente {id: $user_id}) RETURN u", user_id=user_id)
    return 1


def query2(min_price, max_price, stampa):
    result = tx.run("""
        MATCH (p:Prodotto) 
        WHERE p.prezzo >= $min_price AND p.prezzo <= $max_price RETURN p", 
        min_price=min_price, max_price=max_price""")
    return 1


def query3(metodo_pagamento, year, month, day, paese, stampa):
    result = tx.run("""
        MATCH (t:Transazione)
        WHERE t.Metodo_di_Pagamento = $metodo_pagamento and 
        t.Data >= date({year: $year, month: $month, day: $day}) and t.Paese = $paese
        RETURN t
    """, year=year, month=month, day=day, metodo_pagamento=metodo_pagamento, paese=paese)
    return 1


def query4(year, month, day, stampa):
    result = tx.run("""
        MATCH (p:Prodotto)<-[:RIGUARDA]-(t:Transazione)
        WHERE t.Data > date({year: $year, month: $month, day: $day})
        RETURN t
        ORDER BY t.Data, p.Prezzo
    """, year=year, month=month, day=day)
    return 1


def tempo_prima_esecuzione(query_function, *args):
    start_time = timeit.default_timer()
    query_function(*args)
    end_time = timeit.default_timer()
    return end_time - start_time


def tempo_esecuzioni_successive(query_function, *args):
    tempi = [tempo_prima_esecuzione(query_function, *args) for x in range(30)]
    return tempi


def recupera_statistiche(query_function, worksheet, column, stampa, *args):
    row = 2
    tempo = tempo_prima_esecuzione(query_function, *args, stampa)
    tempo_ms = "{:.15f}".format(tempo)
    tempo_ms = float(tempo_ms)*1000
    worksheet.cell(row=row, column=column).value = tempo_ms
    
    row += 1
    tempi = tempo_esecuzioni_successive(query_function, *args, stampa)
    tempo_medio = statistics.mean(tempi)
    tempo_medio_ms = "{:.15f}".format(tempo_medio)
    tempo_medio_ms = float(tempo_medio_ms)*1000
    worksheet.cell(row=row, column=column).value = tempo_medio_ms

    row += 1
    deviazione_standard = statistics.stdev(tempi)
    deviazione_standard_ms = "{:.15f}".format(deviazione_standard)
    deviazione_standard_ms = float(deviazione_standard_ms) * 1000
    worksheet.cell(row=row, column=column).value = deviazione_standard_ms


def esegui_test(file_excel, column, stampa):
    workbook = openpyxl.load_workbook(file_excel)

    worksheet = workbook["query1"]
    recupera_statistiche(query1, worksheet, column, stampa, 383718)

    worksheet = workbook["query2"]
    recupera_statistiche(query2, worksheet, column, stampa, 1, 11000)

    worksheet = workbook["query3"]
    recupera_statistiche(query3, worksheet, column, stampa, "paypal", 2000, 1, 1, "Italy")

    worksheet = workbook["query4"]
    recupera_statistiche(query4, worksheet, column, stampa, 2000, 1, 1)
    

    workbook.save(file_excel)
    workbook.close()


uri = "bolt://localhost:7687"
username = "neo4j"
password = "password"
datasets = ["dataset25", "dataset50", "dataset75", "dataset100"]
file_excel = "excel/risultati2.xlsx"
column = 2

for dataset in datasets:
    graph = Graph(uri, name = dataset, auth=(username, password))
    tx = graph.begin()
    stampa = False
    esegui_test(file_excel, column, stampa)
    column += 2

