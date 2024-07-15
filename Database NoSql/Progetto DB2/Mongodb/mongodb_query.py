from pymongo import MongoClient
from datetime import datetime
import timeit
import statistics
import scipy.stats as stats
import openpyxl


# Connessione al server MongoDB
client = MongoClient("mongodb://root:root@localhost:27017/")

def query1(db, user_id, prima_esecuzione, stampa):
    # Ricerca di un utente specifico
    if (prima_esecuzione):
        db.command({"planCacheClear": "Transazione"})
        db.command({"planCacheClear": "Utente"})
        db.command({"planCacheClear": "Segnalazione_Frode"})
        db.command({"planCacheClear": "Carta_di_Credito"})
        db.command({"planCacheClear": "Prodotto"})

    result = db.Utente.find({"_id": user_id})
    
    if (stampa):
        print(list(result))
    return 1

def query2(db, min_price, max_price, prima_esecuzione, stampa):
    # Ricerca numero prodotti nel range di prezzo specificato
    if (prima_esecuzione):
        db.command({"planCacheClear": "Transazione"})
        db.command({"planCacheClear": "Utente"})
        db.command({"planCacheClear": "Segnalazione_Frode"})
        db.command({"planCacheClear": "Carta_di_Credito"})
        db.command({"planCacheClear": "Prodotto"})

    result = db.Prodotto.find({
        "Prezzo": {
            "$gte": min_price,
            "$lte": max_price
        }
    })
    
    if (stampa):
        print(list(result))
    return 1

def query3(db, metodo_pagamento, year, month, day, paese, prima_esecuzione, stampa):
    # Ricerca transazioni che sono state effettuate con un certo metodo di pagamento, da un certo paese e dopo una certa data
    if (prima_esecuzione):
        db.command({"planCacheClear": "Transazione"})
        db.command({"planCacheClear": "Utente"})
        db.command({"planCacheClear": "Segnalazione_Frode"})
        db.command({"planCacheClear": "Carta_di_Credito"})
        db.command({"planCacheClear": "Prodotto"})

    result = db.Transazione.aggregate([
        {
            "$match": {
                "Metodo_di_Pagamento": metodo_pagamento
            }
        },
        {
            "$match": {
                "Data": {"$gte": datetime(year, month, day)}
            }
        },
        {
            "$match": {
                "Paese": paese
            }
        },
    ])

    if (stampa):
        print(list(result))
    return 1

def query4(db, year, month, day, prima_esecuzione, stampa):
    # Ricerca transazioni che sono state effettuate per acquistare prodotto dopo una certa data in ordine ascendente
    if (prima_esecuzione):
        db.command({"planCacheClear": "Transazione"})
        db.command({"planCacheClear": "Utente"})
        db.command({"planCacheClear": "Segnalazione_Frode"})
        db.command({"planCacheClear": "Carta_di_Credito"})
        db.command({"planCacheClear": "Prodotto"})

    result = db.Transazione.aggregate([
        { "$match": {
                "Data": { "$gt": datetime(year, month, day) }
            }
        },
        { 
            "$lookup": {
                "from": "Prodotto",
                "localField": "Rif_Prodotto",
                "foreignField": "_id",
                "as": "prodotto"
            }
        },
        {
            "$sort": {
                "Data": 1,
                "Prezzo": 1 
            }
        },
        {
            "$project": {
                "Transazione": "$$ROOT"
            }
        }
    ])
    if (stampa):
        print(list(result))
    return 1

def tempo_prima_esecuzione(db, query_function, *args):
    start_time = timeit.default_timer()
    query_function(db, *args)
    end_time = timeit.default_timer()
    return end_time - start_time

def tempo_esecuzioni_successive(db, query_function, *args):
    tempi = [tempo_prima_esecuzione(db, query_function, *args) for x in range(30)]
    return tempi


def recupera_statistiche(db, query_function, worksheet, column, stampa, *args):
    row = 2
    tempo = tempo_prima_esecuzione(db, query_function, *args, True, stampa)
    tempo_ms = "{:.15f}".format(tempo)
    tempo_ms = float(tempo_ms)*1000
    worksheet.cell(row=row, column=column).value = tempo_ms
    
    row += 1
    tempi = tempo_esecuzioni_successive(db, query_function, *args, False, stampa)
    tempo_medio = statistics.mean(tempi)
    tempo_medio_ms = "{:.15f}".format(tempo_medio)
    tempo_medio_ms = float(tempo_medio_ms)*1000
    worksheet.cell(row=row, column=column).value = tempo_medio_ms

    row += 1
    deviazione_standard = statistics.stdev(tempi)
    deviazione_standard_ms = "{:.15f}".format(deviazione_standard)
    deviazione_standard_ms = float(deviazione_standard_ms) * 1000
    worksheet.cell(row=row, column=column).value = deviazione_standard_ms


def esegui_test(dataset, file_excel, column, stampa):
    db = client[dataset]
    workbook = openpyxl.load_workbook(file_excel)

    worksheet = workbook["query1"]
    recupera_statistiche(db, query1, worksheet, column, stampa, 383718)

    worksheet = workbook["query2"]
    recupera_statistiche(db, query2, worksheet, column, stampa, 1, 11000)

    worksheet = workbook["query3"]
    recupera_statistiche(db, query3, worksheet, column, stampa, "paypal", 2000, 1, 1, "Italy")

    worksheet = workbook["query4"]
    recupera_statistiche(db, query4, worksheet, column, stampa, 2000, 1, 1)
    

    workbook.save(file_excel)
    workbook.close()


def main():
    datasets = ["Dataset25", "Dataset50", "Dataset75", "Dataset100"]
    file_excel = "excel/risultati2.xlsx"
    column = 3

    for dataset in datasets:
       # stampa = True if (dataset == "Dataset25") else False
       stampa = False
       esegui_test(dataset, file_excel, column, stampa)
       column += 2


if __name__ == '__main__':
    main()




